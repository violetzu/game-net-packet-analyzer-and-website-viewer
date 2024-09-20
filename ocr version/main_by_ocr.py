import json
import ast
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime, timezone
things = ['']




def read_txt(filename):
    with open(filename, "r", encoding="utf-8") as file:
        return [ast.literal_eval(line.strip()) for line in file]
    
def is_int(variable):
    try:
        int(variable)
        return True
    except ValueError:
        return False

def sort_data(data_dict, sort_index):
    """Sort data based on the specified index."""
    def sorting_key(item):
        value = item[1][sort_index]
        print (item)
        if value:
            if "-" in value and sort_index == 1:
                first_num, second_num = map(int, value.split('-'))
            
                return (first_num, second_num)
            else:
                return (int(value.replace(',', '').replace('-', '')), 0)
        return (0, 0)

    sorted_tuples = sorted(data_dict.items(), key=sorting_key, reverse=True)
    
    # Replace None with "-"
    return [
        (key, (value[0] if value[0] is not None else '-', value[1] if value[1] is not None else '-'))
        for key, value in sorted_tuples
    ]

def populate_sheet(sheet, sorted_data, alliance):
    """Populate the sheet with sorted data and apply formatting."""
    for result in sorted_data:
        name_to_match = result[0].split(' ')[-1].strip()
        alliance_key = next((key for key, value in alliance.items() if name_to_match in value), "")
        sheet.append([result[0], result[1][0], result[1][1], alliance_key])

    for row in sheet.iter_rows(min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='right')

def merge_data(data):
    # Merge data in pairs and return the merged dictionary.
    merged_dict = {}
    for i in range(0, len(data), 2):
        all_keys = set(data[i].keys()) | set(data[i + 1].keys())
        for key in all_keys:
            value1 = data[i].get(key, None)
            value2 = data[i + 1].get(key, None)
            merged_dict[key] = (value1, value2)
    return merged_dict

def set_column_width(sheet):
    # Set column widths for the sheet.
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 11
    sheet.column_dimensions["C"].width = 8
    sheet.column_dimensions["D"].width = 13

def Data_collation(data):
    # Load alliance data
    with open("alliance.json", "r", encoding="utf-8") as f:
        alliance = json.load(f)

    # Update keys
    for player_data in data:
        value = player_data.pop("#7 玩家名字", None)
        if value is not None:
            player_data["#7 Ayo"] = value

    wb = Workbook()
    ws_power_main = wb.active
    ws_power_main.title = "戰力"
    ws_level_main = wb.create_sheet("關卡")

    main_merged_data = merge_data(data)

    # 将元组转换为字典
    formatted_data = {k: {'score': v[0], 'age_range': v[1]} for k, v in main_merged_data.items()}

# 将格式化后的数据保存到JSON文件
    date = datetime.now(timezone.utc).strftime("%m%d")
    with open(f'{date}.json', 'w' , encoding='utf-8') as f:
        json.dump(formatted_data, f, ensure_ascii=False, indent=4)

    
    sorted_by_power_main = sort_data(main_merged_data, 0)
    populate_sheet(ws_power_main, sorted_by_power_main, alliance)

    sorted_by_level_main = sort_data(main_merged_data, 1)
    populate_sheet(ws_level_main, sorted_by_level_main, alliance)

    set_column_width(ws_power_main)
    set_column_width(ws_level_main)

    batch_size = 20
    for idx in range(0, len(data), batch_size):
        batch_data = merge_data(data[idx:idx + batch_size])
        
        sheet_num = idx // batch_size + 1
        ws_power = wb.create_sheet(f"S{sheet_num}戰力")
        ws_level = wb.create_sheet(f"S{sheet_num}關卡")

        sorted_by_power = sort_data(batch_data, 0)
        populate_sheet(ws_power, sorted_by_power, alliance)

        sorted_by_level = sort_data(batch_data, 1)
        populate_sheet(ws_level, sorted_by_level, alliance)

        set_column_width(ws_power)
        set_column_width(ws_level)

    
    wb.save(f'Ranking{date}.xlsx')






def main(raw_data):
    # 幾張圖一個榜
    piture_number=3
    with open("corrections.json", "r", encoding="utf-8") as f:
        corrections = json.load(f)
    sever=0
    data = []
    # 一個迴圈一個榜
    for filenum in range(0, len(raw_data), piture_number):
        results = []

        # 幾張圖一個服
        if filenum%(2*piture_number)==0:
            sever+=1

        # 一個榜幾張圖
        for num in range(piture_number):
            result = raw_data[filenum+num]
            result = [item for item in result if item not in things]
            result.insert(0, "玩家名字")
            
            ids = []
            values = []
            for i, item in enumerate(result):
                if "No." in item:
                    # 向左找 ID
                    id_index = i - 1
                    value_index = i + 1
                    if id_index >= 0:
                        # 如果是戰力補"玩家名字"
                        try:
                            corrected_id_index = corrections.get(result[id_index], result[id_index]).replace(' ', '')
                            if "ower" in corrected_id_index:
                                corrected_id_index = "玩家名字"

                        except ValueError as e:
                            print(f"ValueError: {e}")
                            continue
                        
                        sever_str = sever if sever>9 else f"0{sever}"
                        ids.append(f"#{sever_str} {corrected_id_index}")

                    # 向右找数值
                        def level(result,value_index):
                            value = result[value_index].split(':')[-1].strip().replace('一', '')
    
                            # Case: 'Main Level: 53-46'
                            if "-" in value and is_int(value.split('-')[-1].strip()):
                                value = value.replace(' ', '')
                            
                            # Case: 'Main Level: 53-', '46'
                            elif "-" in value and not is_int(value.split('-')[-1].strip()) and is_int(result[value_index+1]):
                                value = value.replace(' ', '') + result[value_index+1].strip()

                            # Case: 'Main Level:' , '53-46'
                            elif value_index+1 < len(result) and "-" in result[value_index+1] and is_int(result[value_index+1].split('-')[-1].strip()):
                                value = result[value_index+1].replace(' ', '')
                            # Case: 'Main Level: ', '53-', '46'
                            elif value_index+1 < len(result) and "-" in result[value_index+1] and not is_int(result[value_index+1].split('-')[-1].strip()) and is_int(result[value_index+2]):
                                value = result[value_index+1].replace(' ', '') + result[value_index+2].strip()

                            # Case: 'Main Level: ', '53', '46'
                            elif value_index+2 < len(result) and is_int(result[value_index+1]) and is_int(result[value_index+2]):
                                value = result[value_index+1].strip() + "-" + result[value_index+2].strip()

                            # Case: 'Main Level:53 ', '46'
                            elif value_index+1 < len(result) and is_int(result[value_index+1]) :
                                value = value.replace(' ', '') + "-" + result[value_index+1].replace(' ', '')
                            
                            # Case: 'Main Level:53 ', '-46'
                            elif value_index+1 < len(result) and "-" in result[value_index+1]  :
                                value = value.replace(' ', '') + result[value_index+1].replace(' ', '')
                            
                            return value
                    if value_index < len(result):
                        if "Main Level:" in result[value_index]:
                            
                            value = level(result,value_index)

                        elif ("Power:" or "power:") in result[value_index]:
                            # print(result[value_index])
                            value = result[value_index].split(':')[-1].strip() if ":" in result[value_index] else result[value_index]
                            value = value.replace('.',',')
                        else:
                            
                            result[value_index] = "Main Level:"
                            value = level(result,value_index)
                        values.append(value.replace(' ', ''))
        
            # print(ids, values)
            
            results.append(dict(zip(ids, values)))
            
        if piture_number == 3:
            combined_result = {**results[0], **results[1] ,**results[2]}  # 合併n張圖
        elif piture_number == 2:
            combined_result = {**results[0], **results[1]}
        
        print(combined_result)
        data.append(combined_result)
        
    return data



if __name__ == '__main__':
    filename = "raw_data.txt"
    raw_data = read_txt(filename)
    data = main(raw_data)
    Data_collation(data)
